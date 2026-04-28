import pandas as pd
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
import time
import random
from openpyxl import load_workbook
import shutil
import os

import functions as module

def accept_cookies():
    # Fermer le pop-up cookies si présent
    try:
        # XPath : span avec texte "Tout accepter" → bouton parent
        accept_btn = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//span[text()='Tout accepter']/ancestor::button")
            )
        )
        accept_btn.click()
        print("Pop-up cookies accepté")
    except:
        print("Pas de pop-up cookies détecté")

# Charger JSON
dfData = module.load_data("tracks.json")

# Lire le fichier Excel
basepath_ = "Users/Elève/Desktop/developpment"
filepath = f"C:/{basepath_}/frontend/url_list.xlsx"
temp_path = f"C:/{basepath_}/frontend/tmp.xlsx"
# coverDestinationPath = f"C:/{basepath_}/frontend/static/images"
coverDestinationPath = f"E:/Covers"

dfList = pd.read_excel(filepath)

# # Procédure pour rechercher les adresses web de chaque musique sur youtube
options = Options()
options.add_argument("--headless")  # si tu veux en arrière-plan

driver = webdriver.Firefox(options=options)
wait = WebDriverWait(driver, 15)

driver.get("https://www.youtube.com")

# Fermer le pop-up cookies si présent
accept_cookies()

# Temps nécessaire à charger la page d'accueil une fois les cookies acceptés
time.sleep(10)    

# Initilalisation
start_index = 0
i = 0
item = dfData.iloc[i]
artist = item["artist"]
track = item["track"]
ligne = dfList[(dfList['Artist'] == artist) & (dfList['Track'] == track)]

while (not ligne.empty):
    # Récupération des données
    item = dfData.iloc[i]
    artist = item["artist"]
    track = item["track"]

    ligne = dfList[(dfList['Artist'] == artist) & (dfList['Track'] == track)]
    start_index = i
    i += 1

print(f"Dernière chanson enregistrée : n° {start_index}")

for i in range(len(dfData)):

    # Récupération des données
    item = dfData.iloc[i]
    artist = item["artist"]
    track = item["track"]
    album = item["album"]

    ligne = dfList[(dfList['Artist'] == artist) & (dfList['Track'] == track)]

    # Remplir uniquement une ligne non traitée
    if ligne.empty:
        if i % 30 == 0 and i != 0:
            print("")
            print("Redémarrage du navigateur...")
            driver.quit()
            driver = webdriver.Firefox(options=options)
            wait = WebDriverWait(driver, 15)

            driver.get("https://www.youtube.com")
            accept_cookies()
            time.sleep(10) 

        print("")

        query = f"{artist} {track}"
        print("Recherche effectuée :", query)

        # # Recherche
        driver.get("https://www.youtube.com/results?search_query=" + query.replace(" ", "+"))

        # attendre le body
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))

        # Simuler un comportement humain
        time.sleep(random.uniform(2, 5))

        # attendre que la page charge (body présent)
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        # attendre soit des vidéos, soit un état stable
        WebDriverWait(driver, 20).until(
            lambda d: len(d.find_elements(By.CSS_SELECTOR, "a#video-title")) > 0
                    or "Aucun résultat" in d.page_source
        )

        # videos = driver.find_elements(By.CSS_SELECTOR, "a#video-title")
        videos = driver.find_elements(By.CSS_SELECTOR, "ytd-video-renderer")

        if len(videos) == 0:
            print("Aucun résultat trouvé")
            link = None
            dfList.loc[i, 'Downloaded ?'] = 'ERROR'

        else:
            # Choisir la vidéo "provided to youtube by"
            selected_video = None

            for video in videos:
                if (selected_video == None):
                    spans = video.find_elements(By.CSS_SELECTOR, "yt-formatted-string span")

                    for span in spans:
                        description = span.text.lower()

                        if "provided to youtube by" in description:
                            print('Found video "provided to YouTube by" !')
                            selected_video = video

            if (selected_video == None) :
                print('Aucun résultat "provided to YouTube by" trouvé')
                link = None
                dfList.loc[i, 'Downloaded ?'] = 'ERROR'

            else:
                # lien vidéo
                link = selected_video.find_element(By.CSS_SELECTOR, "a#video-title").get_attribute("href")

                # miniature
                result, coverFileName = module.download_cover_from_link(link, artist, album, coverDestinationPath)

                if not(result):
                    print(f"Caractères interdits dans le nom du fichier : {coverFileName}")
                    dfList.loc[i, 'Downloaded ?'] = 'ERROR'
                    print('--------- ERROR ---------')


        # Ajouter l'artiste, le track et le lien dans le dfList
        dfList.loc[i, 'Artist'] = artist
        dfList.loc[i, 'Track'] = track
        dfList.loc[i, 'url'] = link

        # Charger le classeur existant avec openpyxl
        wb = load_workbook(filepath)

        # Écrire avec pandas tout en réutilisant le workbook
        with pd.ExcelWriter(temp_path, engine='openpyxl', mode='a',
                            if_sheet_exists='overlay') as writer:
            dfList.to_excel(writer, sheet_name=wb.sheetnames[0], index=False)

        # Remplace uniquement si succès
        shutil.copy(temp_path, filepath)

        # Avancée globale
        print(f"Avancée globale : {round((i/len(dfData))*100, 2)} %")
        print("Morceaux traités pendant cette exécution :", i + 1- start_index)

driver.quit()