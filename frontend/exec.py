import pandas as pd
import subprocess
import time
from openpyxl import load_workbook
import sys
import os

basepath_ = "Users/vri/Desktop/pythonMp/MaintenancePredictiveConditionnelle"
filepath = f"C:/{basepath_}/3_excel/flotte_valemo.xlsx"

# Lire le fichier Excel
dfParams = pd.read_excel(filepath)

# Liste des parcs pour lesquels on souhaite calculer les données d'apprentissage.
l_parcs = dfParams['NOM_PROJET']
# l_parcs = ['ALBINE']

print(l_parcs)

for parc in l_parcs:
    ligne = dfParams[dfParams['NOM_PROJET'] == parc]

    # Vérifier si une ligne a été trouvée
    if ligne.empty:
        print("")
    else:
        startTime = time.time()

        # Récupérer les valeurs des colonnes
        modele_turbine = ligne['MODELE_TURBINE'].values[0]
        toIgnore = ['OUI', 'CORRIGER', 'INDISPONIBLE']
        if not (ligne['Appris (OUI/NON)'].values[0] in toIgnore):

            if pd.notna(ligne['HL[0]'].values[0]):
                # début = pd.to_datetime(ligne['HL[0]'].values[0]).strftime("%d/%m/%Y")
                # fin = pd.to_datetime(ligne['HL[1]'].values[0]).strftime("%d/%m/%Y")
                début = ligne['HL[0]'].values[0].split(" ")[0]
                fin = ligne['HL[1]'].values[0].split(" ")[0]

                with open(f"C:/{basepath_}/params.txt", "w", encoding="utf-8") as file:
                    file.write(f"{parc} \n")
                    file.write(f"{modele_turbine} \n")
                    file.write(f"{début} \n")
                    file.write(f"{fin} \n")

                print(" ")
                print("parc :", parc)
                print("modèle turbine :", modele_turbine)
                print("début :", début)
                print("fin :", fin)
                print(" ")
                print("  ·----------·  ")

                début = début.replace("/", "-")
                fin = fin.replace("/", "-")
                sourcepath = f"C:/{basepath_}/1_data/11_10min/{parc}_{début}_{fin}.parquet"

                succesful = True
                try:
                    if os.path.exists(sourcepath):
                        print("Utilisation du fichier de prétraitement existant.")
                    else:
                        print("Exécution de pretraitement.py")
                        result1 = subprocess.run([sys.executable, "processing/pretraitement.py"], check=True)

                    print("\n Exécution de main.py")
                    result2 = subprocess.run([sys.executable, "main.py"], check=True)

                except subprocess.CalledProcessError as e:
                    print(
                        f"Le parc {parc} n'a pas été traité : échec lors de l'exécution de {e.cmd}, code retour = {e.returncode}")
                    succesful = False

                except Exception as e:
                    print(f"Le parc {parc} n'a pas été traité pour une autre raison : {e}")
                    succesful = False


                if succesful:
                    output_message = 'Successful'
                    message_excel = 'OUI'


                else:
                    output_message = 'Failed'
                    message_excel = output_message

                print(output_message)


                # Écriture de 'OUI' dans la colonne 'Appris (OUI/NON)' du fichier flotte_valemo.xlsx
                ligne_index = dfParams[dfParams['NOM_PROJET'] == parc].index

                if not ligne_index.empty:
                    # Modifier la valeur dans le DataFrame
                    dfParams.loc[ligne_index, 'Appris (OUI/NON)'] = message_excel

                    # Charger le classeur existant avec openpyxl
                    wb = load_workbook(f"C:/{basepath_}/3_excel/flotte_valemo.xlsx")

                    # Écrire avec pandas tout en réutilisant le workbook
                    with pd.ExcelWriter(f"C:/{basepath_}/3_excel/flotte_valemo.xlsx", engine='openpyxl', mode='a',
                                        if_sheet_exists='overlay') as writer:
                        dfParams.to_excel(writer, sheet_name=wb.sheetnames[0], index=False)



                # Écriture des exécutions et de leur temps d'exécution
                with open(f"C:/{basepath_}/exec.txt", "a", encoding="utf-8") as file:
                    file.write(f"{parc}    -    -    -    {round((time.time() - startTime) / 60, 2)} minutes    -    -    -    {output_message}\n")