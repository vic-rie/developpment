import pandas as pd
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import shutil
import requests

import functions as module


# Charger JSON
dfData = module.load_data("tracks.json")

# Lire le fichier Excel
basepath_ = "Users/Elève/Desktop/developpment"
filepath = f"C:/{basepath_}/frontend/url_list.xlsx"
temp_path = f"C:/{basepath_}/frontend/tmp.xlsx"

dfList = pd.read_excel(filepath)
toIgnore = ['YES', 'ERROR']

# Chemin destination du mp3
# mp3DestinationPath = f"C:/{basepath_}/frontend/static/audio"
mp3DestinationPath = f"E:/Music"

# Ouverture de la page du convertisseur yt -> mp3
options = Options()
options.add_argument("--headless")  # en arrière-plan

driver = webdriver.Firefox(options=options)
wait = WebDriverWait(driver, 15)

driver.get("https://media.ytmp3.gg/fr/")


# Initilalisation
start_index = 0
counter = 0
i = 0
ligne = dfList.iloc[i]

while (ligne["Downloaded ?"] in toIgnore):
    i += 1
    start_index = i
    ligne = dfList.iloc[i]

print("")
print(f"Dernière chanson enregistrée : n° {start_index}")

# Boucle sur les lignes du doc excel 'url_list.xlsx'
for i in range(start_index, len(dfList)):
    # Récupération des infos de la ligne
    ligne = dfList.iloc[i]
    track = ligne["Track"]
    artist = ligne["Artist"]
    url = ligne["url"]

    item = dfData[(dfData["artist"] == artist) & (dfData["track"] == track)]
    album = item["album"].iloc[0]

    if (ligne["Downloaded ?"] not in toIgnore):

        downloaded = False

        input_box = wait.until(
            EC.presence_of_element_located((By.ID, "videoUrl"))
        )

        input_box.clear()
        input_box.send_keys(url)

        convert_button = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//span[text()='Convertir']/ancestor::button")
            )
        )
        convert_button.click()

        download_btn = WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.ID, "download-btn"))
        )

        # attendre que le mp3 soit prêt
        wait.until(lambda d: download_btn.get_attribute("data-url") not in [None, ""])
       
        # récupérer les infos
        mp3_url = download_btn.get_attribute("data-url")
        ytName = download_btn.get_attribute("data-filename")
        duration = int(download_btn.get_attribute("data-duration"))

        rawFilename = f"{artist} - {track}.mp3"
        filename = module.clean_filename(rawFilename)

        print("")
        print(f"Titre : {filename}")
        print(f"Durée de l'audio : {round(duration/60, 2)}min")

        # filtrer
        if (duration >= 1800):
            print(f"Durée supérieure à 30min  -->    audio non téléchargé.")
            dfList.loc[i, 'Downloaded ?'] = 'ERROR' 
            print('--------- ERROR ---------') 
        
        elif (('[' in filename) or ('\\' in filename) or ('/' in filename)
               or ('*' in filename) or ('?' in filename) or (':' in filename)
                or ('"' in filename) or ('<' in filename) or ('>' in filename)
                 or ('|' in filename) or (']' in filename) or (len(filename) == 0)):
            print(f"Caractères interdits dans le nom du fichier : {filename}")
            dfList.loc[i, 'Downloaded ?'] = 'ERROR'
            print('--------- ERROR ---------')   

        else:
            print("Téléchargement...")
            response = requests.get(mp3_url)
            filepath_mp3 = f"{mp3DestinationPath}/{filename}"
            
            try:
                with open(filepath_mp3, "wb") as f:
                    f.write(response.content)
                downloaded = True
                
            except:
                dfList.loc[i, 'Downloaded ?'] = 'ERROR'
                print('ERROR')


            if downloaded:
                module.saveAudio(filepath_mp3, track, artist, album, ytName)

                print("MP3 téléchargé")
                dfList.loc[i, 'Downloaded ?'] = 'YES'

        # Charger le classeur existant avec openpyxl
        wb = load_workbook(filepath)

        # Écrire avec pandas tout en réutilisant le workbook
        with pd.ExcelWriter(temp_path, engine='openpyxl', mode='a',
                            if_sheet_exists='overlay') as writer:
            dfList.to_excel(writer, sheet_name=wb.sheetnames[0], index=False)

        # Remplace uniquement si succès
        shutil.copy(temp_path, filepath)

        # Avancée globale
        counter += 1
        print(f"{i+1} morceaux traités / {len(dfList)} liens disponibles")
        print(f"Avancée globale : {round((i/len(dfData))*100, 2)} %")
        print("Morceaux traités pendant cette exécution :", counter)


        next_btn = WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((By.ID, "next-btn"))
        )
        next_btn.click()

driver.quit()