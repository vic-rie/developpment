import os
import requests
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
import pandas as pd
from openpyxl import load_workbook
import shutil

import functions as module


sp = spotipy.Spotify(
    auth_manager=SpotifyClientCredentials(
        client_id="56212c9c69a14c5a939afde4b66e05f7",
        client_secret="c8d0330fe61a4e72adea93bb9fb5bca1"
    )
)


def clean_filename(name):
    forbidden = '<>:"/\\|?*'
    for char in forbidden:
        name = name.replace(char, "")
    return name.strip()


def download_album_cover(artist, album, destination_folder):
    query = f"artist:{artist} album:{album}"

    results = sp.search(q=query, type="album", limit=1)

    if not results["albums"]["items"]:
        print(f"Aucun album trouvé pour {artist} - {album}")
        return False

    album_data = results["albums"]["items"][0]

    cover_url = album_data["images"][0]["url"]

    filename = clean_filename(f"{artist} - {album}.jpg")
    filepath = os.path.join(destination_folder, filename)

    img_data = requests.get(cover_url).content

    with open(filepath, "wb") as f:
        f.write(img_data)

    print(f"Téléchargé : {filepath}")
    return True


# artist = "Aretha Franklin" 
# album = "Aretha Now"
# coverDestinationPath = f"E:/Covers_from_spotify"
# download_album_cover(artist, album, coverDestinationPath)

# print("ok")

# Charger JSON
dfData = module.load_data("tracks.json")

# Lire le fichier Excel
basepath_ = "Users/Elève/Desktop/developpment"
filepath = f"C:/{basepath_}/frontend/url_list.xlsx"
temp_path = f"C:/{basepath_}/frontend/tmp.xlsx"
coverDestinationPath = f"E:/Covers_from_spotify"

dfList = pd.read_excel(filepath)


# Initilalisation
start_index = 0
i = 0
item = dfData.iloc[i]
artist = item["artist"]
track = item["track"]
ligne = dfList[(dfList['Artist'] == artist) & (dfList['Track'] == track)]


while (((ligne["Cover"] == 'YES').iloc[0]) or (ligne["Cover"] == 'NOT FOUND').iloc[0]) and (i < len(dfData)):
    item = dfData.iloc[i]
    artist = item["artist"]
    track = item["track"]

    ligne = dfList[(dfList['Artist'] == artist) & (dfList['Track'] == track)]
    start_index = i
    i += 1

print(f"Dernière pochette téléchargée : n° {start_index}")

for i in range(len(dfData)):
    # Récupération des données
    item = dfData.iloc[i]
    artist = item["artist"]
    track = item["track"]
    album = item["album"]

    ligne = dfList[(dfList['Artist'] == artist) & (dfList['Track'] == track)]
    indexLigne = ligne.index[0]

    # Remplir uniquement une ligne non traitée
    if (((ligne["Cover"] != 'YES').iloc[0]) or ((ligne["Cover"] != 'NOT FOUND').iloc[0])):
        print("")

        coverName = clean_filename(f"{artist} - {album}.jpg")
        coverFullPath = os.path.join(coverDestinationPath, coverName)

        if not(os.path.exists(coverFullPath)):

            query = f"{artist} {track}"
            print("Requête effectuée sur Spotify :", query)

            result = download_album_cover(artist, album, coverDestinationPath)

            if result:
                dfList.loc[indexLigne, "Cover"] = 'YES'
            else:
                dfList.loc[indexLigne, "Cover"] = 'NOT FOUND'

        else:
            print("Pochette déjà téléchargée")
            dfList.loc[indexLigne, "Cover"] = 'YES'


        # Ajouter l'artiste, le track et le lien dans le dfList
        print(f"{artist} - {album}")

        
    if (i % 10 == 0):
        # Charger le classeur existant avec openpyxl
        wb = load_workbook(filepath)

        # Écrire avec pandas tout en réutilisant le workbook
        with pd.ExcelWriter(temp_path, engine='openpyxl', mode='a',
                            if_sheet_exists='overlay') as writer:
            dfList.to_excel(writer, sheet_name=wb.sheetnames[0], index=False)

        # Remplace uniquement si succès
        shutil.copy(temp_path, filepath)

        # Avancée globale
        print(f"Avancée globale : {round((i/len(dfData))*100, 2)} %")
        print("Morceaux traités pendant cette exécution :", i + 1 - start_index)